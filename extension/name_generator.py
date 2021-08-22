from openpyxl import load_workbook
import random

name = []
word = []
number = []
seperator = []

def maxRow(i):
    wb = load_workbook(filename="./data/data.xlsx",data_only=True)
    data = wb["name"]
    count = 1
    while True:
        value = data.cell(row=count+1,column=i).value
        if value != None:
            count += 1
        else:
            return count
        

def createLists():
    wb = load_workbook(filename="./data/data.xlsx")
    data = wb["name"]
    i = 1
    while i <= 4:
        max = maxRow(i)
        for row in data.iter_rows(min_row=2,max_row=max,min_col=i,max_col=i,values_only=True):
            row = str(row)
            row = row.replace("(","")
            row = row.replace(",)","")
            if i == 1:
                name.append(row)
            elif i == 2:
                word.append(row)
            elif i == 3:
                number.append(row)
            elif i == 4:
                seperator.append(row)
            else:
                raise ValueError("row not found")
        i += 1

def buildName(n):
    createLists()
    for _ in range(n):
        one = random.choice(name)
        two = random.choice(word)
        #num = random.choice(number)
        num = random.randint(1,1000)
        sep = random.choice(seperator)
        #first_step = Seperator(one,sep)
        first_step = one + str(sep)
        second_step = first_step + two
        final_step = second_step + str(num)
        final_step = final_step.replace("'","")
        return final_step, num

def Seperator(txt,sep):
    first_step = list(txt)
    first_step.remove("'")
    first_step.remove("'")
    position = random.randint(1,len(first_step))
    first_step.insert(position,sep)
    first_word = ""
    for letter in first_step:
        first_word = first_word + letter
    first_word =  first_word.replace("'","")
    return first_word

def createEmail(number):
    main = 'instabotnetwork1337'
    ending = '@gmail.com'
    mail_name = main + '.' + str(number)
    mail = main + '.' + str(number) + ending
    return mail, mail_name


name, num = buildName(1)
mail, mailname = createEmail(num)
pw = "test.python"
print("",name,"\n",mailname,"\n",mail,"\n",pw)