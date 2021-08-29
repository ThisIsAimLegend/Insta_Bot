from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
import mariadb
import sys
from sys import exit
import keyring

#Account that should be added
new_bot = str("Doublixah")
pw = str("test.python")
email = str("test.mail.trichter@gmail.com")


def  check_memory():
    wb = load_workbook(filename="./data/data.xlsx")
    w_bot = wb["Bots"]
    for row in w_bot.iter_cols(min_row=1, min_col=1, max_row=1, values_only=True):
        if new_bot in row:
            exit("Bot already in bot memory!")

def bot_memory():
    wb = load_workbook(filename="./data/data.xlsx")
    w_bot = wb["Bots"]
    new_col = w_bot.max_column
    new_col += 1
    bot_num = w_bot.cell(row=1,column=new_col,value=new_bot).border = Border(bottom=Side(style="thick"),left=Side(style="thick"))
    wb.save("./data/data.xlsx")
    print("New bot memory created:",new_bot)


def check_account():
    wb = load_workbook(filename="./data/data.xlsx")
    w_bot = wb["Account-Daten"]
    for row in w_bot.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        if new_bot in row:
            exit("Bot already in account list!")

def bot_account():
    wb = load_workbook(filename="./data/data.xlsx")
    w_bot = wb["Account-Daten"]
    new_row = w_bot.max_row
    new_row += 1
    w_bot.cell(row=new_row, column=1, value=new_bot)
    w_bot.cell(row=new_row, column=2, value=pw).border = Border(left=Side(style="thick"))
    w_bot.cell(row=new_row, column=3, value=email).border = Border(left=Side(style="thick"))
    wb.save("./data/data.xlsx")
    print("New bot account in list:",new_bot)

def excel():
    check_account()
    bot_account()
    check_memory()
    bot_memory()

class BotAction:
    def __init__(self):
        try:
            mariadb.connect(
                user="root",
                password=keyring.get_password("databank","root"),
                host=keyring.get_password("IP","database"),
                port=3306,
                database="instagram_bot"
            )
        except mariadb.Error as e:
            print(f"Error connecting to MariaDB Platform: {e}")
            sys.exit(1)

        connection = mariadb.connect(
            user="root",
            password=keyring.get_password("databank","root"),
            host=keyring.get_password("IP","database"),
            port=3306,
            database="instagram_bot"
        )
        cur = connection.cursor()

        self.connection = connection
        self.cur = cur

    def addBot(self):
        try:
            self.cur.execute("INSERT INTO account_information (username,password,email) VALUES (%s,%s,%s)",(new_bot,pw,email))
            print("Added Bot to databank")

        except:
            print("Something went wrong")
        self.connection.close()

#Log = BotAction()
#Log.addBot()

def Bot():
    connection = mariadb.connect(
        user="root",
        password=keyring.get_password("databank","root"),
        host=keyring.get_password("IP","database"),
        port=3306,
        database="instabot"
    )
    cur = connection.cursor()

    #cur.execute("INSERT INTO account_information (username,password,email) VALUES (?,?,?)",(new_bot,pw,email))
    #print("Added ",new_bot," to the databank")
    #connection.commit()
    connection.close()

Bot()
