from openpyxl import load_workbook, Workbook

#Account that should be deleted
search_object = "aha lul"


wb = load_workbook(filename="./data/data.xlsx")
w_accounts = wb["Account-Daten"]
i = 1
for row in w_accounts.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
    i += 1
    if search_object in row:
        account = row
        print(account)
        w_accounts.delete_rows(i,1)
        print("Deleted Account in row:",i)
wb.save("./data/data.xlsx")
wb.close()

wb = load_workbook(filename="./data/data.xlsx")
w_accounts = wb["Bots"]
i = 0
for column in w_accounts.iter_cols(min_col=1, min_row=1, max_row=1, values_only=True):
    i += 1
    if search_object in column:
        account = column
        print(account)
        w_accounts.delete_cols(i)
        print("Deleted Account in column:",i)
wb.save("./data/data.xlsx")