from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from sys import exit
import random


def getNumberofAccounts():
    wb = load_workbook(filename="./data/data.xlsx")
    w_bot = wb["Account-Daten"]
    maximum = w_bot.max_row
    maximum -= 1
    wb.close()
    return maximum

def getAllAccounts():
    wb = load_workbook(filename="./data/data.xlsx")
    w_bot = wb["Account-Daten"]
    maximum = w_bot.max_row
    i = 1
    for acc in w_bot.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        print(i,":",acc)
        i += 1
    wb.close()

def getAccount(acc):
    wb = load_workbook(filename="./data/data.xlsx")
    w_bot = wb["Account-Daten"]
    maximum = w_bot.max_row
    bot = int(acc + 1)
    if bot <= maximum:
        name = w_bot.cell(row=bot,column=1).value
        pw = w_bot.cell(row=bot,column=2).value
        #print(name,pw)
        wb.close()
        return name, pw
    else:
        raise ModuleNotFoundError("This bot doesn't exist!")


def SearchForAccount(target,bot):
    wb = load_workbook(filename="./data/data.xlsx")
    w_bot = wb["Bots"]
    for column in w_bot.iter_cols(min_row=2, min_col=bot, max_col=bot, values_only=True):
        if target in column:
            return True
        else:
            return False
    wb.close()

def AddTargetToMemory(target,bot):
    wb = load_workbook(filename="./data/data.xlsx")
    w_bot = wb["Bots"]
    i = 2
    for account in w_bot.iter_cols(min_row=2, min_col=bot, max_col=bot, values_only=True):
        for element in account:
            if element != None:
                i += 1
    w_bot.cell(row=i, column=bot, value=target).border = Border(left=Side(style="thick"))
    wb.save("data.xlsx")
    wb.close()
    print("Successfully added",target,"in line",i,"to the bot memory")
