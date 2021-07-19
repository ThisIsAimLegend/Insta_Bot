import datetime as dt
import excel_actions as ea
from openpyxl import load_workbook, Workbook

timer = dt.datetime.now()
timer = timer.strftime("[%H:%M:%S %d.%m.%Y]")

def getNumberofAccounts():
    wb = load_workbook(filename="data.xlsx")
    w_bot = wb["Account-Daten"]
    maximum = w_bot.max_row
    maximum -= 1
    return maximum

def chooseAccounts(count,max_accounts):
    bot_accounts = []
    if count <= max_accounts:
        for i in range(count):
            n = i + 1
            bot_accounts.append(n)
    else:
        raise RuntimeError("Not enough existing bot accounts!")
    return bot_accounts

chooseAccounts(3,getNumberofAccounts()))
